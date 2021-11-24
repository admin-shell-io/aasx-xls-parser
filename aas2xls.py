'''

The MIT License (MIT)

Copyright (c) 2021 Nestfield Co., Ltd. 
<https://www.nestfield.co.kr>             
Author: Wonseok Song

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
THE SOFTWARE.


The openpyxl is under the MIT/Expat license
(see https://github.com/chronossc/openpyxl/blob/master/LICENCE).

'''

import sys
import openpyxl as xl
import json
from openpyxl.styles import Alignment

PARSER_VERSION_STRING = '2021.10.07.build-1'


# chohpower
#if len(sys.argv) != 2:
#    print("usage: xls2aas [xlsx file name]")
if len(sys.argv) != 4:
    print("usage  : aas2xls [aas file name] [output name] [result file name]")
    print("example: aas2xls aas.json output.xlsx result.txt")
    sys.exit()


#-------------------- count variables ------------------------------
count_asset = 0
count_asset_invalid = 0

count_aas = 0
count_aas_invalid = 0
count_aas_no_submodel = 0

count_submodel = 0
count_submodel_invalid = 0
count_submodel_unmatch = 0

count_property = 0
count_property_no_value_type = 0
count_property_no_value = 0
count_property_cd_unmatch = 0
count_property_no_preferred_name = 0
count_property_no_short_name = 0
count_property_no_definition = 0
count_property_invalid = 0

count_collection = 0
count_collection_invalid = 0

count_concept_description = 0

def outValidataionResult(fp):
    fp.write("\n-------------------- validataion report -------------------")
    fp.write('\nnumber of asset : ok........{}'.format(count_asset))
    fp.write('\nnumber of asset : invalid...{}'.format(count_asset_invalid))
    fp.write("\n")

    fp.write('\nnumber of asset-administration-shell : ok.........................{}'.format(count_aas))
    fp.write('\nnumber of asset-administration-shell : ok (no submodel defined)...{}'.format(count_aas_no_submodel))
    fp.write('\nnumber of asset-administration-shell : invalid....................{}'.format(count_aas_invalid))
    fp.write("\n")

    fp.write('\nnumber of submodel : ok...........................................................{}'.format(count_submodel))
    fp.write('\nnumber of submodel : ok (defined in AAS, but actual information is not defined)...{}'.format(count_submodel_unmatch))
    fp.write('\nnumber of submodel : invalid......................................................{}'.format(count_submodel_invalid))
    fp.write("\n")

    fp.write('\nnumber of property : ok...........................................................{}'.format(count_property))
    fp.write('\nnumber of property : ok (value-type is not assigned)..............................{}'.format(count_property_no_value_type))
    fp.write('\nnumber of property : ok (initial-value is not assigned)...........................{}'.format(count_property_no_value))
    fp.write('\nnumber of property : ok (concept-description is not matched)......................{}'.format(count_property_cd_unmatch))
    fp.write('\nnumber of property : ok (preferred-name is not assigned in concept-description)...{}'.format(count_property_no_preferred_name))
    fp.write('\nnumber of property : ok (short-name is not assigned in concept-description).......{}'.format(count_property_no_short_name))
    fp.write('\nnumber of property : ok (definition is not assigned in concept-description).......{}'.format(count_property_no_definition))
    fp.write('\nnumber of property : invalid......................................................{}'.format(count_property_invalid))
    fp.write("\n")

    fp.write('\nnumber of submodel-element-collection : ok........{}'.format(count_collection))
    fp.write('\nnumber of submodel-element-collection : invalid...{}'.format(count_collection_invalid))
    fp.write("\n")

    fp.write('\nnumber of concept-description : ok...{}'.format(count_concept_description))

    fp.write("\n")
    return

#-------------------- common sub functions --------------------------
VTYPE_STRING            = 'string'
VTYPE_LANG_STRING       = 'langString'

fp_result = open(sys.argv[3], mode = 'wt')


def outMessage(msg):
    print(msg)
    fp_result.write(msg)
    fp_result.write("\n")
    return



COLUMN_ASSET                    = 0     # 'A'
COLUMN_AAS_LEVEL0               = 1     # 'B'
COLUMN_AAS_LEVEL1               = 2     # 'C'
COLUMN_AAS_LEVEL2               = 3     # 'C'
COLUMN_SUBMODEL                 = 4     # 'E'
COLUMN_COLLECTION_LEVEL0        = 5     # 'F'
COLUMN_COLLECTION_LEVEL1        = 6     # 'G'
COLUMN_COLLECTION_LEVEL2        = 7     # 'H'
COLUMN_COLLECTION_LEVEL3        = 8     # 'I'
COLUMN_COLLECTION_LEVEL4        = 9     # 'J'
COLUMN_COLLECTION_LEVEL5        = 10    # 'K'
COLUMN_FIELD_NAME               = 11    # 'L'
COLUMN_PROPERTY                 = 12    # 'M'
COLUMN_OPTIONS                  = 13    # 'N'
COLUMN_ASSET_AAS_SM_ID_IRI      = 14    # 'O'
COLUMN_REFERENCE_TYPE           = 15    # 'P'
COLUMN_REFERENCE_LOCAL          = 16    # 'Q'
COLUMN_SEMANTICS_NAME           = 17    # 'R'
COLUMN_SEMANTICS_SHORT_NAME     = 18    # 'S'
COLUMN_SEMANTICS_PREF_NAME      = 19    # 'T'
COLUMN_SEMANTICS_DATA_TYPE      = 20    # 'U'
COLUMN_SEMANTICS_IRI            = 21    # 'V'
COLUMN_SEMANTICS_IRDI           = 22    # 'W'
COLUMN_INITIAL_VALUE            = 23    # 'X'
COLUMN_ARRAY                    = 24    # 'Y'
COLUMN_ENGINEERING_UNIT         = 25    # 'Z'
COLUMN_PROPERTY_VALUE_TYPE      = 26    # 'AA'
COLUMN_SEMANTICS_DEFINITION     = 27    # 'AB'
COLUMN_FIELD_TAG_NAME           = 28    # 'AC'
COLUMN_NOTE                     = 29    # 'AD'

MAX_EXCEL_COLUMNS               = 30


SMETYPE_PROPERTY            = 0     # Submodel element : Property
SMETYPE_MLP                 = 1     # Submodel element : Multi-language-property
SMETYPE_COLLECTION          = 2     # Submodel element : collection
SMETYPE_FILE                = 3     # Submodel element : File
SMETYPE_REF		    = 4     # Submodel element : ReferenceEmelent

excelRows = []

def getDictItem(userDict, findingKey):
    for key in userDict.keys():
        if key.lower() == findingKey.lower():
            return userDict[key]

    return None

def getDictItem_depth2(userDict, findingKey1, findingKey2):

    findingDict = None

    for key in userDict.keys():
        if key.lower() == findingKey1.lower():
            findingDict = userDict[key]
            break

    if findingDict == None:
        return None

    for key in findingDict:
        if key.lower() == findingKey2.lower():
            return findingDict[key]
    
    return None

def getDictItem_depth3(userDict, findingKey1, findingKey2, findingKey3):

    findingDict = None

    for key in userDict.keys():
        if key.lower() == findingKey1.lower():
            findingDict = userDict[key]
            break

    if findingDict == None:
        return None

    for key in findingDict:
        if key.lower() == findingKey2.lower():
            findingDict = findingDict[key]
            break

    if findingDict == None:
        return None

    for key in findingDict:
        if key.lower() == findingKey3.lower():
            return findingDict[key]
    
    return None

def mlValueToString(mlv):
    # parsing 'preferredName'
    if mlv == None:
        return None

    stringArray = []
    for langStr in mlv:
        stringArray.append('@{}:{}'.format(getDictItem(langStr, 'language'), getDictItem(langStr, 'text') ))

    return '\n'.join(stringArray)


def keyValueToString(kv_array):
    # parsing 'keys [{type/local/value/index/idType} ... {}]'
    if len(kv_array) < 1:
        return None
	
    str_list = []
	
    for kv in kv_array:
        t  = getDictItem(kv, 'type')
        l = getDictItem(kv, 'local')
        v = getDictItem(kv, 'value')
        i = getDictItem(kv, 'index')
        d = getDictItem(kv, 'idType')
		
        if t == None or l == None or v == None or i == None or d == None:
            return None
			
        str_list.append('{{ {},{},{},{},{} }}'.format(i, t, l, d, v))
	
    return ''.join(str_list)

def getAdminShellFor(listAdminShells, assetId):

    for shell in listAdminShells:
        item_asset = getDictItem(shell, 'asset')
        if item_asset == None:
            continue;

        item_keys = getDictItem(item_asset, 'keys')
        if item_keys == None:
            continue;

        if len(item_keys) < 1:
            continue

        item_value = getDictItem(item_keys[0], 'value')
        if item_value == None:
            continue

        if item_value == assetId:
            return shell
    
    
    return None


def getSubmodel(allSubmodels, smInAdminShell):

    # get 'submodel id' from submodel information in Administration shell
    item_keys = getDictItem(smInAdminShell, 'keys')
    if item_keys == None:
        return None

    if len(item_keys) < 1:
        return None

    findingSmId = getDictItem(item_keys[0], 'value')
    if findingSmId == None:
        return None


    # find actual submodel dictionary
    for sm in allSubmodels:
        identification = getDictItem(sm, 'identification')
        if identification == None:
            continue;
    
        smId = getDictItem(identification, 'id')
        if smId == None:
            continue

        if smId == findingSmId:
            return sm

    return None


def getConceptDescription(allConcepts, conceptId):

    # find actual submodel dictionary
    for concept in allConcepts:
        identification = getDictItem(concept, 'identification')
        if identification == None:
            continue;
    
        cdId        = getDictItem(identification, 'id')
        cdIdType    = getDictItem(identification, 'idType')

        if cdId == None or cdIdType == None:
            continue

        if cdId == conceptId:
            return concept, cdIdType

    return None, None



def getSMElementType(sme):
    modelType = getDictItem(sme, 'modelType')
    if modelType == None:
        return None

    name = getDictItem(modelType, 'name')
    if name == None:
        return None

    if name.lower() == 'property':
        return SMETYPE_PROPERTY

    elif name.lower() == 'multilanguageproperty':
        return SMETYPE_MLP

    elif name.lower() == 'submodelelementcollection':
        return SMETYPE_COLLECTION

    elif name.lower() == 'file':
        return SMETYPE_FILE
		
    elif name.lower() == 'referenceelement':
        return SMETYPE_REF

    return None



def writeAsset(asset):
    row = [None for i in range(MAX_EXCEL_COLUMNS)]

    # parsing 'idShort'
    idShort = getDictItem(asset, 'idShort')
    if idShort == None:
        outMessage('error  : some asset has no "idShort"')
        return None

    row[COLUMN_ASSET] = idShort


    # parsing 'identification'
    identification = getDictItem(asset, 'identification')
    if identification == None:
        outMessage('error  : asset "{}" has no "identification"'.format(idShort))
        return None


    assetId = getDictItem(identification, 'id')
    if assetId == None:
        outMessage('error  : asset "{}" has no "id"'.format(idShort))
        return None

    assetIdType = getDictItem(identification, 'idType')
    if assetIdType.lower() != 'iri':
        outMessage('warning: asset "{}" has "id" but idType is not IRI'.format(idShort))

    row[COLUMN_ASSET_AAS_SM_ID_IRI] = assetId

    # add to excel rows
    #excelRows.append(row)
    return row


def writeAdminShell(shell):
    row = [None for i in range(MAX_EXCEL_COLUMNS)]
    
    # parsing 'idShort'
    idShort = getDictItem(shell, 'idShort')
    if idShort == None:
        outMessage('error  : some adminShell has no "idShort"')
        return None

    row[COLUMN_AAS_LEVEL0] = idShort

    # parsing 'identification'
    adminShellId = getDictItem_depth2(shell, 'identification', 'id')
    if adminShellId == None:
        outMessage('error  : adminShell "{}" has no "id"'.format(idShort))
        return None

    row[COLUMN_ASSET_AAS_SM_ID_IRI] = adminShellId

    # parsing 'asset - keys - type/local'
    assetKeys = getDictItem_depth2(shell, 'asset', 'keys')
    if assetKeys == None:
        outMessage('error  : invalid asset-keys in adminShell "{}"'.format(idShort))
        return None

    if len(assetKeys) < 1:
        outMessage('error  : no asset-keys  in adminShell "{}"'.format(idShort))
        return None

    row[COLUMN_REFERENCE_TYPE]  = getDictItem(assetKeys[0], 'type')
    row[COLUMN_REFERENCE_LOCAL] = getDictItem(assetKeys[0], 'local')

    #excelRows.append(row)
    return row


def writeSubmodel(sm):
    row = [None for i in range(MAX_EXCEL_COLUMNS)]

    # parsing 'idShort'
    idShort = getDictItem(sm, 'idShort')
    if idShort == None:
        outMessage('error  : some submodel has no "idShort"')
        return None

    row[COLUMN_SUBMODEL] = idShort

    # parsing 'identification'
    submodelId = getDictItem_depth2(sm, 'identification', 'id')
    if submodelId == None:
        outMessage('error  : Submodel "{}" has no "id"'.format(idShort))
        return None

    row[COLUMN_ASSET_AAS_SM_ID_IRI] = submodelId
    
    
    # parsing 'output string'
    optionStringList = []

    optCategory = getDictItem(sm, 'category')
    if optCategory != None:
        optionStringList.append('category={}'.format(optCategory))

    optKind = getDictItem(sm, 'kind')
    if optKind != None:
        optionStringList.append('kind={}'.format(optKind))

    if len(optionStringList) > 0:
        optionString = ','.join(optionStringList)
        row[COLUMN_OPTIONS] = optionString

    # parsing 'semanticId - keys - type/local/value'
    semanticKeys = getDictItem_depth2(sm, 'semanticId', 'keys')
    if semanticKeys == None:
        outMessage('error  : invalid semanticId-keys in Submodel "{}"'.format(idShort))
        return None

    if len(semanticKeys) < 1:
        # outMessage('error  : no semanticId-keys  in Submodel "{}"'.format(idShort))    # change error to warning
        # return None
        outMessage('warning: no semanticId-keys  in Submodel "{}"'.format(idShort))

    else:
        row[COLUMN_REFERENCE_TYPE]  = getDictItem(semanticKeys[0], 'type')
        row[COLUMN_REFERENCE_LOCAL] = getDictItem(semanticKeys[0], 'local')

        semanticIdType = getDictItem(semanticKeys[0], 'idType')
        if semanticIdType == None:
            outMessage('error  : semanticId-keys in Submodel "{}" has no "idType"'.format(idShort))
            return None

        if semanticIdType.lower() == 'iri':
            row[COLUMN_SEMANTICS_IRI]   = getDictItem(semanticKeys[0], 'value')
        else:
            row[COLUMN_SEMANTICS_IRDI]  = semanticIdType

        if (row[COLUMN_SEMANTICS_IRI] == None) and (row[COLUMN_SEMANTICS_IRDI] == None):
            outMessage('error  : semanticId-keys in Submodel "{}" has no "value"'.format(idShort))
            return None


    #excelRows.append(row)
    return row


def writeProperty(prop, smeType, conceptDictionary):
    global count_property
    global count_property_no_value_type
    global count_property_no_value
    global count_property_cd_unmatch
    global count_property_no_preferred_name
    global count_property_no_short_name
    global count_property_no_definition
    global count_property_invalid

    
    row = [None for i in range(MAX_EXCEL_COLUMNS)]

    # parsing 'idShort'
    idShort = getDictItem(prop, 'idShort')
    if idShort == None:
        outMessage('error  : some property has no "idShort"')
        return None

    if smeType == SMETYPE_PROPERTY:
        row[COLUMN_PROPERTY] = 'Prop:' + idShort

    elif smeType == SMETYPE_MLP:
        row[COLUMN_PROPERTY] = 'MLP:' + idShort

    elif smeType == SMETYPE_FILE:
        row[COLUMN_PROPERTY] = 'File:' + idShort
		
    elif smeType == SMETYPE_REF:
        row[COLUMN_PROPERTY] = 'Ref:' + idShort



    # parsing 'option string'
    optionStringList = []

    optCategory = getDictItem(prop, 'category')
    if optCategory != None:
        optionStringList.append('category={}'.format(optCategory))

    optKind = getDictItem(prop, 'kind')
    if optKind != None:
        optionStringList.append('kind={}'.format(optKind))

    if smeType == SMETYPE_FILE:
        optMimeType = getDictItem(prop, 'mimeType')
        if optMimeType != None:
            if len(optMimeType) > 0:
                optionStringList.append('mimeType={}'.format(optMimeType))

    if len(optionStringList) > 0:
        optionString = ','.join(optionStringList)
        row[COLUMN_OPTIONS] = optionString

    # parsing 'valueType'
    if smeType == SMETYPE_PROPERTY or smeType == SMETYPE_MLP:
        row[COLUMN_PROPERTY_VALUE_TYPE] = getDictItem_depth3(prop, 'valueType', 'dataObjectType', 'name' )
        if row[COLUMN_PROPERTY_VALUE_TYPE] == None:
            outMessage('warning: valueType is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
            count_property_no_value_type += 1
     
    # parsing 'initial value'
    valueObject = getDictItem(prop, 'value')
    if valueObject != None:
        if isinstance(valueObject, dict):
            
            # check if value is type of 'langString'
            if smeType == SMETYPE_MLP:
                langStringValue = getDictItem(valueObject, 'langString')
                if langStringValue != None:
                    row[COLUMN_INITIAL_VALUE] = mlValueToString(langStringValue)
					
            elif smeType == SMETYPE_REF:
                refKeyValueArray = getDictItem(valueObject, 'keys')
                if refKeyValueArray != None:
                    row[COLUMN_INITIAL_VALUE] = keyValueToString(refKeyValueArray)
        
        else:
            row[COLUMN_INITIAL_VALUE] = valueObject
    else:
        outMessage('warning: value is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
        count_property_no_value += 1
     

        

    # parsing 'semanticId - keys - type/local/value' (for finding ConceptDescription)
    # change 'semanticId' from Mandatory to Optional
    semanticKeys = getDictItem_depth2(prop, 'semanticId', 'keys')
    if semanticKeys == None:
        #outMessage('error  : invalid semanticId-keys in Property "{}"'.format(idShort))
        #return None
        outMessage('warning: invalid semanticId-keys in Property "{}"'.format(idShort))
        count_property_cd_unmatch += 1
        return row

    # change 'semanticId' from Mandatory to Optional
    if len(semanticKeys) < 1:
        #outMessage('error  : no semanticId-keys  in Property "{}"'.format(idShort))
        #return None
        outMessage('warning: no semanticId-keys  in "{}"'.format(row[COLUMN_PROPERTY]))
        count_property_cd_unmatch += 1
        return row


    row[COLUMN_REFERENCE_TYPE]  = getDictItem(semanticKeys[0], 'type')
    row[COLUMN_REFERENCE_LOCAL] = getDictItem(semanticKeys[0], 'local')

    semanticIdType = getDictItem(semanticKeys[0], 'idType')
    if semanticIdType == None:
        outMessage('warning: semanticId-keys in "{}" has no "idType"'.format(row[COLUMN_PROPERTY]))
        count_property_cd_unmatch += 1
        return row

    conceptId = getDictItem(semanticKeys[0], 'value')
    if conceptId == None:
        outMessage('warning: semanticId-keys in "{}" has no "value"'.format(row[COLUMN_PROPERTY]))
        count_property_cd_unmatch += 1
        return row

    # find ConceptDescription 
    cd, cdIdType = getConceptDescription(conceptDictionary, conceptId)
    if cd != None and cdIdType != None:
        if cdIdType.lower() == 'iri':
            row[COLUMN_SEMANTICS_IRI] = conceptId

        elif cdIdType.lower() == 'irdi':
            row[COLUMN_SEMANTICS_IRDI] = conceptId

        else:
            outMessage('warning: ConceptDescription idType is not IRI/IRDI in "{}"'.format(row[COLUMN_PROPERTY]))

        cdIdShort = getDictItem(cd, 'idShort')
        if cdIdShort == None:
            outMessage('warning: ConceptDescription has no "idShort" in "{}"'.format(row[COLUMN_PROPERTY]))
            return row
        else:
            row[COLUMN_SEMANTICS_NAME] = cdIdShort
        
        embeddedDataSpecs = getDictItem(cd, 'embeddedDataSpecifications')
        if embeddedDataSpecs != None:

            if len(embeddedDataSpecs) >= 1:

                # parsing 'preferredName'
                prefName = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'preferredName')
                if prefName != None:
                    row[COLUMN_SEMANTICS_PREF_NAME] = mlValueToString(prefName)

                else:
                    outMessage('warning: "PreferredName" of ConceptDescription is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
                    count_property_no_preferred_name += 1

                # parsing 'shortName'
                shortName = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'shortName')
                if shortName != None:
                    row[COLUMN_SEMANTICS_SHORT_NAME] = mlValueToString(shortName)

                else:
                    outMessage('warning: "shortName" of ConceptDescription is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
                    count_property_no_short_name += 1

                # parsing 'definition'
                definition = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'definition')
                if definition != None:
                    row[COLUMN_SEMANTICS_DEFINITION] = mlValueToString(definition)

                else:
                    outMessage('warning: "definition" of ConceptDescription is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
                    count_property_no_definition += 1


                # parsing 'unit' and 'dataType'
                row[COLUMN_ENGINEERING_UNIT]    = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'unit')
                row[COLUMN_SEMANTICS_DATA_TYPE] = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'dataType')

            else:
                outMessage('warning: ConceptDescription has no valid "embeddedDataSpecifications" in "{}"'.format(row[COLUMN_PROPERTY]))
                count_property_cd_unmatch += 1

        else:
            outMessage('warning: ConceptDescription is not specified in "{}"'.format(row[COLUMN_PROPERTY]))
            count_property_cd_unmatch += 1
     
    else:
        outMessage('warning: ConceptDescription has no "embeddedDataSpecifications" in "{}"'.format(row[COLUMN_PROPERTY]))
        count_property_cd_unmatch += 1
     

    #excelRows.append(row)
    return row


def writeSMECollection(collection, depth, conceptDictionary):
    row = [None for i in range(MAX_EXCEL_COLUMNS)]
    
    # parsing 'idShort'
    idShort = getDictItem(collection, 'idShort')
    if idShort == None:
        outMessage('error  : some SMC has no "idShort"')
        return None

    if depth == 0:
        row[COLUMN_COLLECTION_LEVEL0] = idShort

    elif depth == 1:
        row[COLUMN_COLLECTION_LEVEL1] = idShort
    
    elif depth == 2:
        row[COLUMN_COLLECTION_LEVEL2] = idShort

    elif depth == 3:
        row[COLUMN_COLLECTION_LEVEL3] = idShort

    elif depth == 4:
        row[COLUMN_COLLECTION_LEVEL4] = idShort
		
    elif depth == 5:
        row[COLUMN_COLLECTION_LEVEL5] = idShort

    else:
        outMessage('error  : SMC "{}" level is too deep'.format(idShort))
        return None



    # parsing 'option string'
    optionStringList = []

    optCategory = getDictItem(collection, 'category')
    if optCategory != None:
        optionStringList.append('category={}'.format(optCategory))

    optKind = getDictItem(collection, 'kind')
    if optKind != None:
        optionStringList.append('kind={}'.format(optKind))

    optOrdered = getDictItem(collection, 'ordered')
    if optOrdered != None:
        optionStringList.append('ordered={}'.format(optOrdered))

    optAllowDup = getDictItem(collection, 'allowDuplicates')
    if optAllowDup != None:
        optionStringList.append('allowDuplicates={}'.format(optAllowDup))

    if len(optionStringList) > 0:
        optionString = ','.join(optionStringList)
        row[COLUMN_OPTIONS] = optionString


    # parsing 'semanticId - keys - type/local/value' (for finding ConceptDescription)
    # change 'semanticId' from Mandatory to Optional
    semanticKeys = getDictItem_depth2(collection, 'semanticId', 'keys')
    if semanticKeys == None:
        #outMessage('error  : invalid semanticId-keys in SMC "{}"'.format(idShort))
        #return None
        outMessage('warning: invalid semanticId-keys in SMC "{}"'.format(idShort))
        return row


    if len(semanticKeys) < 1:
        #outMessage('error  : no semanticId-keys  in SMC "{}"'.format(idShort))
        #return None
        outMessage('warning: no semanticId-keys  in SMC "{}"'.format(idShort))
        return row


    row[COLUMN_REFERENCE_TYPE]  = getDictItem(semanticKeys[0], 'type')
    row[COLUMN_REFERENCE_LOCAL] = getDictItem(semanticKeys[0], 'local')

    semanticIdType = getDictItem(semanticKeys[0], 'idType')
    if semanticIdType == None:
        outMessage('warning: semanticId-keys in SMC "{}" has no "idType"'.format(idShort))
        return row

    conceptId = getDictItem(semanticKeys[0], 'value')
    if conceptId == None:
        outMessage('warning: semanticId-keys in SMC "{}" has no "value"'.format(idShort))
        return row

    # find ConceptDescription 
    cd, cdIdType = getConceptDescription(conceptDictionary, conceptId)
    if cd != None and cdIdType != None:
        if cdIdType.lower() == 'iri':
            row[COLUMN_SEMANTICS_IRI] = conceptId

        elif cdIdType.lower() == 'irdi':
            row[COLUMN_SEMANTICS_IRDI] = conceptId

        else:
            outMessage('warning: ConceptDescription idType is not IRI/IRDI in SMC "{}"'.format(idShort))

        cdIdShort = getDictItem(cd, 'idShort')
        if cdIdShort == None:
            outMessage('warning: ConceptDescription has no "idShort" in SMC "{}"'.format(idShort))
            return row
        else:
            row[COLUMN_SEMANTICS_NAME] = cdIdShort
        

        embeddedDataSpecs = getDictItem(cd, 'embeddedDataSpecifications')
        if embeddedDataSpecs != None:

            if len(embeddedDataSpecs) >= 1:

                # parsing 'preferredName'
                prefName = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'preferredName')
                if prefName != None:
                    row[COLUMN_SEMANTICS_PREF_NAME] = mlValueToString(prefName)

                else:
                    outMessage('warning: "PreferredName" of ConceptDescription is not specified in SMC "{}"'.format(idShort))

                # parsing 'shortName'
                shortName = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'shortName')
                if shortName != None:
                    row[COLUMN_SEMANTICS_SHORT_NAME] = mlValueToString(shortName)

                else:
                    outMessage('warning: "PreferredName" of ConceptDescription is not specified in SMC "{}"'.format(idShort))
        
                # parsing 'definition'
                definition = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'definition')
                if definition != None:
                    row[COLUMN_SEMANTICS_DEFINITION] = mlValueToString(definition)

                else:
                    outMessage('warning: "definition" of ConceptDescription is not specified in SMC "{}"'.format(idShort))


                # parsing 'unit' and 'dataType'
                row[COLUMN_ENGINEERING_UNIT]    = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'unit')
                row[COLUMN_SEMANTICS_DATA_TYPE] = getDictItem_depth2(embeddedDataSpecs[0], 'dataSpecificationContent', 'dataType')

            else:
                outMessage('warning: ConceptDescription has no valid "embeddedDataSpecifications" in SMC "{}"'.format(idShort))

        else:
            outMessage('warning: ConceptDescription is not specified in SMC "{}"'.format(idShort))
     
    else:
        outMessage('warning: ConceptDescription has no "embeddedDataSpecifications" in SMC "{}"'.format(idShort))
     

    #excelRows.append(row)
    return row


def writeSME(sme, depth, rowBase, conceptDictionary):

    global count_property
    global count_property_invalid
    global count_collection
    global count_collection_invalid


    smeType = getSMElementType(sme)
    if smeType == None:
        return rowBase

    row = None

    if smeType == SMETYPE_PROPERTY or smeType == SMETYPE_MLP or smeType == SMETYPE_FILE or smeType == SMETYPE_REF:
        row = writeProperty(sme, smeType, conceptDictionary)
        if row != None:
            count_property += 1
            if depth == 0:
                excelRows.insert(rowBase, row)
                rowBase += 1
            else:
                excelRows.append(row)
        else:
            count_property_invalid += 1

    elif smeType == SMETYPE_COLLECTION:
        row = writeSMECollection(sme, depth, conceptDictionary)
        if row != None:
            count_collection += 1
            excelRows.append(row)

            subElements = getDictItem(sme, 'value')
            if subElements != None:
                for element in subElements:
                    writeSME(element, depth + 1, rowBase, conceptDictionary)
        else:
            count_collection_invalid += 1

    return rowBase


#-------------------- making aas.xlsx codes -------------------------
adminShells = []
assets = []
submodels = []
conceptDescriptions = []

outMessage('Parser version : {}'.format(PARSER_VERSION_STRING))


with open(sys.argv[1], mode= 'rt') as aas_file:
    aasx = json.load(aas_file)

    # get main dictionaries
    aasx_keys = aasx.keys()
    for aasx_key in aasx_keys:
        if aasx_key.lower() == 'assetadministrationshells':
            adminShells = aasx[aasx_key]

        elif aasx_key.lower() == 'assets':
            assets = aasx[aasx_key]

        elif aasx_key.lower() == 'submodels':
            submodels = aasx[aasx_key]

        elif aasx_key.lower() == 'conceptdescriptions':
            conceptDescriptions = aasx[aasx_key]

    count_concept_description = len(conceptDescriptions)

    # fill excel file
    for asset in assets:
        row_asset = writeAsset(asset)
        if row_asset == None:
            count_asset_invalid += 1
            continue

        else:
            excelRows.append(row_asset)
            count_asset += 1

        #find 'adminShell' linked to this 'asset'
        shell = getAdminShellFor(adminShells, row_asset[COLUMN_ASSET_AAS_SM_ID_IRI])
        if shell == None:
            outMessage('warning: no adminShell for asset "{}" is defined'.format(row_asset[COLUMN_ASSET]))
            count_asset_invalid += 1
            continue

        row_shell = writeAdminShell(shell)
        if row_shell == None:
            count_aas_invalid += 1
            continue

        else:
            excelRows.append(row_shell)
            count_aas += 1

        listSubmodel = getDictItem(shell, 'submodels')
        if listSubmodel == None:
            count_aas_no_submodel += 1
            continue

        # writing 'Submodel' of selected 'AdminShell'
        for smInShell in listSubmodel:
            sm = getSubmodel(submodels, smInShell)
            if sm == None:
                count_submodel_unmatch += 1
                continue

            row_sm = writeSubmodel(sm)
            if row_sm == None:
                count_submodel_invalid += 1
                continue
            else:
                excelRows.append(row_sm)
                count_submodel += 1

            smc_depth   = 0 # submodel collection depth
            row_begin   = len(excelRows)

            # writing 'Submodel Elements' oc selected 'Submodel'
            listSMElements = getDictItem(sm, 'submodelElements')
            if listSMElements == None:
                continue

            for sme in listSMElements:
                row_begin = writeSME(sme, smc_depth, row_begin, conceptDescriptions)



# write output excel file
if len(excelRows) > 0:
    #xls_wb      = Workbook()
#chohpower
    xls_wb      = xl.load_workbook('/opt/bin/aas-design-template.xlsx')
    xls_sheet   = xls_wb.active

    line = 3
    for aasElement in excelRows:
        for col in range(1, MAX_EXCEL_COLUMNS + 1, 1) :
            editCell = xls_sheet.cell(row=line, column=col)
            editCell.value = aasElement[col - 1]

        line += 1

    xls_wb.save(sys.argv[2])
    xls_wb.close()

outMessage('Completed...')
outValidataionResult(fp_result)

# close result file
fp_result.close()

# make result aasx.json file
#chohpower
#with open(sys.argv[2], 'w') as fp:
#    json.dump(aasx, fp, ensure_ascii = False, indent = 4)


